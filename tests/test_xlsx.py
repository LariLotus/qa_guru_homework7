from openpyxl import load_workbook
import os

# TODO оформить в тест, добавить ассерты и использовать универсальный путь
from constants import RESOURCES_PATH


def test_xlsx():
    xlsx_file = os.path.join(RESOURCES_PATH, 'file_example_XLSX_50.xlsx')
    workbook = load_workbook(xlsx_file)
    sheet = workbook.active
    print(sheet.cell(row=3, column=2).value)

    assert sheet.cell(row=3, column=2).value == 'Mara'
